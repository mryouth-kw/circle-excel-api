from fastapi import FastAPI, UploadFile, File, Form, HTTPException
from fastapi.responses import StreamingResponse
from fastapi.middleware.cors import CORSMiddleware

import openpyxl
import pandas as pd
import zipfile
import tempfile
import os
import csv
import requests
import traceback
import shutil
import gc

from io import BytesIO

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=False,
    allow_methods=["*"],
    allow_headers=["*"],
)

columns_to_search = [
    "患者ID（文字列型）",
    "患者ID"
]

CSV_URL = (
    "https://docs.google.com/spreadsheets/d/"
    "1Ys9iA8aIqyAgnOoUrnu1CZT6adT2E12hSuDmaH4blXQ/"
    "export?format=csv"
)

# =========================
# cache
# =========================
mapping_cache = None


# =========================
# normalize
# =========================
def normalize(text):

    if text is None:
        return ""

    text = (
        str(text)
        .replace("\ufeff", "")
        .replace("\u3000", " ")
        .replace("\r", "")
        .replace("\n", "")
        .strip()
    )

    try:

        num = float(text)

        if num.is_integer():
            return str(int(num))

    except:
        pass

    return text


# =========================
# safe filename
# =========================
def safe_filename(text):

    return (
        str(text)
        .replace("/", "_")
        .replace("\\", "_")
        .replace(":", "_")
        .replace("*", "_")
        .replace("?", "_")
        .replace('"', "_")
        .replace("<", "_")
        .replace(">", "_")
        .replace("|", "_")
    )


# =========================
# load mapping
# =========================
def load_mapping():

    global mapping_cache

    if mapping_cache is not None:
        return mapping_cache

    print("LOAD GOOGLE SHEET")

    try:

        res = requests.get(
            CSV_URL,
            timeout=15
        )

        res.raise_for_status()

    except Exception as e:

        print("GOOGLE SHEET ERROR:", str(e))

        raise HTTPException(
            status_code=500,
            detail=f"Google Sheets load failed: {str(e)}"
        )

    res.encoding = "utf-8"

    reader = csv.reader(
        res.text.splitlines()
    )

    mapping_cache = {}

    for row in reader:

        if len(row) < 2:
            continue

        key = normalize(row[0])
        value = normalize(row[1])

        if key:
            mapping_cache[key] = value

    print(
        f"MAPPING COUNT: {len(mapping_cache)}"
    )

    return mapping_cache


def get_target_value(circle_id):

    mapping = load_mapping()

    return mapping.get(
        normalize(circle_id)
    )


# =========================
# target sheet
# =========================
def get_target_sheet(wb):

    sheetnames = wb.sheetnames

    if len(sheetnames) == 0:
        return None

    first_sheet = sheetnames[0]

    print("FIRST SHEET:", first_sheet)

    if normalize(first_sheet) == "検索情報":

        print("検索情報 sheet detected")

        if len(sheetnames) >= 2:

            print(
                "USE SECOND SHEET:",
                sheetnames[1]
            )

            return wb[sheetnames[1]]

        return None

    print("USE FIRST SHEET")

    return wb[first_sheet]


# =========================
# xls -> xlsx
# =========================
def convert_xls_to_xlsx(
    xls_path,
    xlsx_path
):

    excel_file = pd.ExcelFile(
        xls_path,
        engine="xlrd"
    )

    with pd.ExcelWriter(
        xlsx_path,
        engine="openpyxl"
    ) as writer:

        for sheet_name in excel_file.sheet_names:

            print(
                "CONVERT SHEET:",
                sheet_name
            )

            df = pd.read_excel(
                xls_path,
                sheet_name=sheet_name,
                engine="xlrd",
                header=None
            )

            df.to_excel(
                writer,
                sheet_name=sheet_name,
                header=False,
                index=False
            )

    del excel_file
    gc.collect()


@app.get("/")
@app.head("/")
def root():
    return {"message": "API OK"}


@app.post("/process")
async def process_excel(
    files: list[UploadFile] = File(...),
    circle_id: str = Form(...),
    visit: str = Form(...)
):

    print("=" * 80)
    print("PROCESS START")
    print("=" * 80)

    target_value = get_target_value(circle_id)

    print("CIRCLE ID:", circle_id)
    print("TARGET VALUE:", target_value)

    if not target_value:

        raise HTTPException(
            status_code=404,
            detail=f"{circle_id} に対応する値が見つかりません"
        )

    temp_dir = tempfile.mkdtemp()

    zip_path = os.path.join(
        temp_dir,
        f"調査票2_{circle_id}_Visit-{visit}.zip"
    )

    try:

        with zipfile.ZipFile(
            zip_path,
            mode="w",
            compression=zipfile.ZIP_DEFLATED,
            compresslevel=1
        ) as zipf:

            for file in files:

                wb = None

                try:

                    print("=" * 80)
                    print("FILE:", file.filename)
                    print("=" * 80)

                    safe_input_name = safe_filename(
                        file.filename
                    )

                    input_path = os.path.join(
                        temp_dir,
                        safe_input_name
                    )

                    # =========================
                    # upload save
                    # =========================
                    with open(input_path, "wb") as f:

                        while True:

                            chunk = await file.read(1024 * 1024)

                            if not chunk:
                                break

                            f.write(chunk)

                    await file.close()

                    base_name, ext = os.path.splitext(
                        file.filename
                    )

                    ext = ext.lower()

                    safe_target_value = safe_filename(
                        target_value
                    )

                    # =========================
                    # xls
                    # =========================
                    if ext == ".xls":

                        print("XLS MODE")

                        converted_path = os.path.join(
                            temp_dir,
                            f"converted_{base_name}.xlsx"
                        )

                        convert_xls_to_xlsx(
                            input_path,
                            converted_path
                        )

                        wb = openpyxl.load_workbook(
                            converted_path,
                            data_only=False
                        )

                        output_filename = (
                            f"{safe_target_value}_{base_name}.xlsx"
                        )

                        output_path = os.path.join(
                            temp_dir,
                            output_filename
                        )

                    # =========================
                    # xlsx / xlsm
                    # =========================
                    else:

                        print("XLSX/XLSM MODE")

                        wb = openpyxl.load_workbook(
                            input_path,
                            keep_vba=(ext == ".xlsm"),
                            data_only=False
                        )

                        output_filename = (
                            f"{safe_target_value}_{safe_input_name}"
                        )

                        output_path = os.path.join(
                            temp_dir,
                            output_filename
                        )

                    print(
                        "SHEETS:",
                        wb.sheetnames
                    )

                    # =========================
                    # target sheet
                    # =========================
                    ws = get_target_sheet(wb)

                    if ws is None:

                        print("TARGET SHEET NOT FOUND")

                        wb.save(output_path)

                        wb.close()

                        zipf.write(
                            output_path,
                            arcname=output_filename
                        )

                        continue

                    print(
                        "TARGET SHEET:",
                        ws.title
                    )

                    target_col_index = None
                    header_row_index = None

                    # =========================
                    # header search
                    # =========================
                    for row in ws.iter_rows(
                        min_row=1,
                        max_row=min(10, ws.max_row),
                        values_only=True
                    ):

                        headers = [
                            normalize(v)
                            for v in row
                        ]

                        print(
                            "HEADER ROW:",
                            headers
                        )

                        for col_name in columns_to_search:

                            if col_name in headers:

                                target_col_index = (
                                    headers.index(col_name) + 1
                                )

                                header_row_index = (
                                    headers.index(headers[0]) + 1
                                )

                                print(
                                    "FOUND COLUMN:",
                                    col_name
                                )

                                break

                        if target_col_index:
                            break

                    if header_row_index is None:
                        header_row_index = 1

                    print(
                        "TARGET COLUMN INDEX:",
                        target_col_index
                    )

                    # =========================
                    # column not found
                    # =========================
                    if not target_col_index:

                        print("COLUMN NOT FOUND")

                        wb.save(output_path)

                        wb.close()

                        zipf.write(
                            output_path,
                            arcname=output_filename
                        )

                        continue

                    matched_count = 0

                    rows_to_keep = []

                    # =========================
                    # keep header
                    # =========================
                    for row in ws.iter_rows(
                        min_row=1,
                        max_row=header_row_index,
                        values_only=True
                    ):

                        rows_to_keep.append(
                            list(row)
                        )

                    # =========================
                    # scan rows
                    # =========================
                    for row in ws.iter_rows(
                        min_row=header_row_index + 1,
                        values_only=True
                    ):

                        raw_value = row[
                            target_col_index - 1
                        ]

                        cell_value = normalize(
                            raw_value
                        )

                        if cell_value == target_value:

                            matched_count += 1

                            rows_to_keep.append(
                                list(row)
                            )

                    print(
                        "MATCHED COUNT:",
                        matched_count
                    )

                    print(
                        "ROWS TO KEEP:",
                        len(rows_to_keep)
                    )

                    # =========================
                    # recreate workbook
                    # =========================
                    new_wb = openpyxl.Workbook()

                    new_ws = new_wb.active
                    new_ws.title = ws.title

                    for row_values in rows_to_keep:

                        new_ws.append(row_values)

                    print(
                        "SAVE:",
                        output_filename
                    )

                    new_wb.save(output_path)

                    new_wb.close()

                    wb.close()

                    del wb
                    del new_wb
                    del rows_to_keep

                    gc.collect()

                    zipf.write(
                        output_path,
                        arcname=output_filename
                    )

                    # =========================
                    # cleanup per file
                    # =========================
                    try:
                        os.remove(input_path)
                    except:
                        pass

                    try:
                        os.remove(output_path)
                    except:
                        pass

                    print(
                        "ZIP ADD:",
                        output_filename
                    )

                except Exception as e:

                    print("=" * 80)
                    print("ERROR OCCURRED")
                    print(traceback.format_exc())
                    print("=" * 80)

                    if wb:
                        try:
                            wb.close()
                        except:
                            pass

                    raise HTTPException(
                        status_code=500,
                        detail=f"{file.filename}: {str(e)}"
                    )

                finally:

                    gc.collect()

        # =========================
        # zip response
        # =========================
        return StreamingResponse(
            open(zip_path, "rb"),
            media_type="application/zip",
            headers={
                "Content-Disposition":
                f'attachment; filename="{os.path.basename(zip_path)}"'
            }
        )

    finally:

        gc.collect()

        try:
            shutil.rmtree(temp_dir)
        except:
            pass